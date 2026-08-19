#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate QA test case files (TestRail CSV + tracking CSV + XLSX) dari tabel
test case di `QA Test Plan - Accounting.md`.

Output:
  qa-test-cases-testrail.csv  → format import TestRail (Case ID, Title, Section,
                                Type, Priority, Estimate, References,
                                Preconditions, Steps, Expected)
  qa-test-cases-tracker.csv   → spreadsheet pelacakan (UTF-8 BOM, Excel-friendly)
                              (kolom Status/Tanggal Run/Environment + Pembuat
                              Test + Link Bug)
  qa-test-cases.xlsx          → workbook berformat (filter, freeze, ringkasan)
  qa-test-results-testrail.csv      → template HASIL eksekusi per run, siap
                                      di-import balik ke TestRail sebagai
                                      result (kolom Status/Comment/Elapsed/…)
  qa-test-results-testrail.json     → payload add_results_for_cases, ditulis
                                      OTOMATIS oleh konverter setelah
                                      regenerasi bila ada status terisi
                                      (kosong → INFO, JSON tidak ditulis)
  qa-test-results-template.xlsx     → worksheet eksekusi per run: dropdown
                                      Passed/Failed/Blocked, ringkasan otomatis
                                      (COUNTIF) + release gate S1

Status TestRail (default): 1=Passed, 2=Blocked, 3=Untested (default,
TIDAK bisa dikirim sebagai hasil), 4=Retest, 5=Failed. Nama status di CSV
impor harus persis: Passed, Failed, Blocked, Retest, Skipped, Untested.

Status, Tanggal Run, dan Environment yang SUDAH diisi di output sebelumnya
(`qa-test-cases.xlsx`, fallback `qa-test-cases-tracker.csv`) dipertahankan
per baris saat regenerasi — hanya test case baru / kolom yang masih kosong
yang jatuh ke default ('Not Run', RUN_DATE, ENVIRONMENT). Jadi QA bisa mengisi
hasil eksekusi per baris lalu menjalankan ulang generator tanpa kehilangan data.

`--sample`: hasilkan CONTOH run dengan status acak-deterministik ke file
`qa-test-cases-sample-tracker.csv`, `qa-test-cases-sample.xlsx`,
`qa-test-results-sample-testrail.csv`, dan `qa-test-results-sample.json`
(payload add_results_for_cases) sebagai referensi format import — file asli
TIDAK disentuh. Status memakai seed TETAP (SAMPLE_SEED) yang dicetak di
output — run ulang menghasilkan status yang sama persis (reproduksi
membutuhkan QA_RUN_DATE/QA_ENVIRONMENT yang sama pula, lihat header & QA
Test Plan §8.1).

Menjalankan:  python scripts/generate-qa-test-cases.py [--sample]
"""
from __future__ import annotations

import argparse
import csv
import importlib.util
import json
import os
import re
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
PLAN = ROOT / "QA Test Plan - Accounting.md"

# Nilai default kolom otomatis run (bisa diedit per baris di Excel).
# Prioritas override: argumen CLI (--run-date / --environment) > env
# (QA_RUN_DATE / QA_ENVIRONMENT) > default. Env QA_RUN_DATE dipakai oleh
# scripts/check-qa-sync.py agar regenerasi di CI tidak dianggap berubah hanya
# karena ganti tanggal.
RUN_DATE_DEFAULT = os.environ.get("QA_RUN_DATE") or date.today().isoformat()
ENVIRONMENT_DEFAULT = os.environ.get("QA_ENVIRONMENT") or "QA Local (mock API)"
# Saat --run-date di-override via CLI, paksa semua baris memakai tanggal baru
# (bukan hanya baris kosong). Ini memastikan check-qa-sync tidak gagal karena
# perbedaan tanggal antara tracker on-disk dan hasil regenerasi.
_OVERRIDE_RUN_DATE = False
# Nilai default kolom "Pembuat Test" (bisa diedit per baris di Excel).
TEST_AUTHOR_DEFAULT = "Tim QA"

# Baseline jumlah test case di QA Test Plan. Ini acuan lunak (warning, bukan
# error): penambahan/pengurangan TC atau RG di plan TIDAK boleh memaksa edit
# kode — generator tetap jalan, cukup menampilkan peringatan agar drift jumlah
# terlihat dan baseline di sini bisa diperbarui bila perubahan disengaja.
EXPECTED_TC = 137
EXPECTED_RG = 15

# Status yang memicu peringatan S1
OPEN_STATUSES = {"Not Run", "Fail"}


def _update_baseline(new_tc: int, new_rg: int) -> None:
    """Perbarui EXPECTED_TC dan EXPECTED_RG di source code file ini.

    Membaca file sendiri (Path(__file__)), mencari baris
    `EXPECTED_TC = N` dan `EXPECTED_RG = N`, mengganti angkanya,
    lalu menulis ulang. Efektif — cukup dipanggil sekali saat
    --update-baseline dikombinasikan dengan perubahan QA Test Plan.
    """
    src = Path(__file__)
    content = src.read_text(encoding="utf-8")
    import re
    new_content = re.sub(
        r'EXPECTED_TC = \d+',
        f'EXPECTED_TC = {new_tc}',
        content,
    )
    new_content = re.sub(
        r'EXPECTED_RG = \d+',
        f'EXPECTED_RG = {new_rg}',
        new_content,
    )
    if new_content != content:
        src.write_text(new_content, encoding="utf-8")
        print(f"[OK] EXPECTED_TC = {new_tc}, EXPECTED_RG = {new_rg} — baseline diperbarui di {src.name}")
    else:
        print(f"[OK] Baseline sudah sesuai (TC={new_tc}, RG={new_rg}) — tidak ada perubahan")

# Epoch timestamp untuk byte-determinism XLSX — openpyxl selalu menulis
# dcterms:modified di core.xml dengan datetime.now() saat wb.save(), jadi
# kita perlu post-process ZIP agar timestamp-nya tetap stabil antar run.
_XLSX_EPOCH = "2026-01-01T00:00:00Z"


def _pin_xlsx_modified(path: Path) -> None:
    """Normalize dcterms:modified in docProps/core.xml so the XLSX is
    byte-deterministic across runs (openpyxl uses wall-clock time)."""
    import re as _re
    import zipfile as _zf

    with _zf.ZipFile(path, "r") as zin:
        core = zin.read("docProps/core.xml").decode("utf-8")
        patched = _re.sub(
            r"<dcterms:modified[^>]*>[^<]+</dcterms:modified>",
            f"<dcterms:modified xmlns:dcterms=\"http://purl.org/dc/terms/\" "
            f"xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\" "
            f"xsi:type=\"dcterms:W3CDTF\">{_XLSX_EPOCH}</dcterms:modified>",
            core,
        )
        items = []
        for info in zin.infolist():
            data = (
                patched.encode("utf-8")
                if info.filename == "docProps/core.xml"
                else zin.read(info.filename)
            )
            items.append((info, data))

    tmp = Path(str(path) + ".pin")
    with _zf.ZipFile(tmp, "w", _zf.ZIP_DEFLATED) as zout:
        for info, data in items:
            info.date_time = (2026, 1, 1, 0, 0, 0)  # pin ZIP header timestamps
            zout.writestr(info, data)
    tmp.replace(path)


# ---------------------------------------------------------------------------
# Peta story (AAJ-xxx) → sprint, dari matriks traceability (QA Test Plan §5)
# ---------------------------------------------------------------------------
STORY_SPRINT: dict[str, int] = {}
for sprint, n_range in {
    1: range(1, 10),
    2: range(10, 16),
    3: range(16, 20),
    4: range(20, 25),
    5: range(25, 30),
    6: range(30, 35),
}.items():
    for n in n_range:
        STORY_SPRINT[f"AAJ-{n:03d}"] = sprint

SEVERITY_PRIORITY = {
    "S1": "1 - Critical",
    "S2": "2 - High",
    "S3": "3 - Medium",
    "S4": "4 - Low",
}

DEFAULT_PRECOND = "Reset data seed (mock API di-reset); login sebagai Rina (admin)"

# Preconditions khusus per test case (awalan ID)
PRECOND_OVERRIDES: dict[str, str] = {
    "TC-COA-03": "COA kosong (belum dimuat template)",
    "TC-COA-12": "COA kosong (belum dimuat template)",
    "TC-COA-13": "COA sudah terisi akun",
    "TC-PER-03": "Reset data seed; Maret 2026 masih punya 2 jurnal draft",
    "TC-PER-04": "Reset data seed; Maret 2026 masih punya 2 jurnal draft",
    "TC-PER-05": "Reset data seed; Maret 2026 sudah ditutup",
    "TC-PER-06": "Reset data seed; periode Maret 2026 sudah ditutup",
    "TC-JRN-22": "Reset data seed; periode Februari 2026 tertutup",
    "TC-APR-04": "Login sebagai Dimas (accountant) — tanpa izin approve",
    "TC-RLE-02": "Login sebagai Budi (viewer)",
    "TC-RLE-03": "Login sebagai Dimas (accountant)",
    "TC-RLE-04": "Login bergantian sebagai Rina/Dimas/Budi",
    "TC-RLE-05": "Login sebagai akuntan; entitas CV Karya Mandiri (ent-002) tersedia",
    "TC-RLE-06": "Login sebagai akuntan; data ent-001 sudah terisi",
    "TC-STT-02": "Server mock API dimatikan",
    "TC-PRF-01": "Seed 10.000 jurnal (POST /admin/seed-bulk)",
    "TC-PRF-02": "Seed 10.000 jurnal (POST /admin/seed-bulk)",
    "TC-PRF-03": "Seed 10.000 jurnal (POST /admin/seed-bulk)",
    "TC-BNK-01": "Integrasi bank nonaktif (feature flag OFF)",
    "TC-BNK-02": "Integrasi bank aktif (partner terhubung)",
    "TC-BNK-03": "Integrasi bank aktif; ada jurnal existing untuk rekonsiliasi",
    "RG-": "State seed; mock API + prototipe berjalan (localhost:4000 / :5173)",
}

# Modul spesifik per awalan ID (untuk section yang berisi banyak modul)
ID_MODULE: dict[str, str] = {
    "TC-SRC": "Pencarian Global",
    "TC-RVS": "Reverse",
    "TC-ATT": "Lampiran",
    "TC-APR": "Approval",
    "TC-ONB": "Onboarding",
    "TC-BNK": "Bank",
}

# ---------------------------------------------------------------------------
# Parsing markdown
# ---------------------------------------------------------------------------
def clean(cell: str) -> str:
    """Bersihkan sel markdown: backtick, **bold**, whitespace ekstra."""
    cell = cell.replace("`", "").replace("**", "")
    return re.sub(r"\s+", " ", cell).strip()


def header_section(line: str) -> str | None:
    """Ekstrak nama section dari header markdown (## atau ###)."""
    m = re.match(r"^##+\s+\d+\.\d+\s+(.+?)(?:\s*\(.*\))?$", line)
    if m:
        return clean(re.sub(r"—.*$", "", m.group(1)))
    m = re.match(r"^##+\s+\d+\.\s+(.+?)(?:\s*\(.*\))?$", line)
    if m:
        return clean(m.group(1))
    m = re.match(r"^###\s+(Arus Kas)", line)
    if m:
        return m.group(1)
    return None


def parse_tables() -> list[dict]:
    lines = PLAN.read_text(encoding="utf-8").splitlines()
    section = "Umum"
    records: list[dict] = []
    i = 0
    n = len(lines)

    while i < n:
        line = lines[i].strip()
        sec = header_section(line)
        if sec:
            section = sec
            i += 1
            continue
        if not line.startswith("|"):
            i += 1
            continue

        # Kumpulkan baris tabel
        table: list[list[str]] = []
        while i < n and lines[i].strip().startswith("|"):
            cells = [c.strip() for c in lines[i].strip().strip("|").split("|")]
            # Abaikan baris pemisah (---)
            if not all(re.fullmatch(r":?-{3,}:?", c) for c in cells if c.strip()):
                table.append(cells)
            i += 1

        if len(table) < 2:
            continue

        header = [clean(c) for c in table[0]]
        try:
            idx_id = header.index("ID")
            if "Test Case" in header:
                idx_title, idx_steps, idx_exp, idx_ref, idx_sev = (
                    header.index("Test Case"), header.index("Langkah"),
                    header.index("Hasil Diharapkan"), header.index("AC"),
                    header.index("Sev"),
                )
            elif "Skenario" in header:  # tabel RG-01..12
                idx_title, idx_steps, idx_exp = (
                    header.index("Skenario"), header.index("Langkah Inti"),
                    header.index("Verifikasi Kunci"),
                )
                idx_ref, idx_sev = None, None
            else:
                continue
        except ValueError:
            continue

        for row in table[1:]:
            if len(row) <= idx_id:
                continue
            cid = clean(row[idx_id])
            if not re.match(r"^(TC-|RG-)", cid):
                continue
            title = clean(row[idx_title])
            steps = clean(row[idx_steps])
            expected = clean(row[idx_exp])
            ref = clean(row[idx_ref]) if idx_ref is not None and idx_ref < len(row) else ""
            sev = clean(row[idx_sev]) if idx_sev is not None and idx_sev < len(row) else ""
            records.append({
                "id": cid,
                "title": title,
                "steps": steps,
                "expected": expected,
                "ref": ref,
                "sev": sev,
                "section": section,
            })
    return records


# ---------------------------------------------------------------------------
# Enrichment
# ---------------------------------------------------------------------------
def story_of(ref: str) -> str:
    m = re.search(r"AAJ-\d{3}", ref)
    return m.group(0) if m else ""


def sprint_of(ref: str) -> str:
    s = story_of(ref)
    return f"Sprint {STORY_SPRINT[s]}" if s in STORY_SPRINT else ""


def module_of(rec: dict) -> str:
    for prefix, mod in ID_MODULE.items():
        if rec["id"].startswith(prefix):
            return mod
    return rec["section"]


def precond_of(rec: dict) -> str:
    for prefix, text in PRECOND_OVERRIDES.items():
        if rec["id"].startswith(prefix):
            return text
    return DEFAULT_PRECOND


def load_existing_metadata() -> dict[str, dict[str, str]]:
    """Baca Status / Tanggal Run / Environment yang sudah diisi QA dari output
    sebelumnya, agar regenerasi TIDAK me-reset hasil eksekusi ke default
    (Status → 'Not Run', Tanggal Run/Environment → nilai default run).

    Prioritas: `qa-test-cases.xlsx` (sheet 'Test Cases') — file paling baru
    karena ditulis setelah tracker CSV. Fallback: `qa-test-cases-tracker.csv`
    (mis. XLSX belum pernah dibuat / korup). Nilai sel kosong diabaikan
    (kolom yang kosong jatuh ke default saat enrich).
    """
    existing: dict[str, dict[str, str]] = {}

    xlsx = ROOT / "qa-test-cases.xlsx"
    if xlsx.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx, read_only=True, data_only=True)
            ws = wb["Test Cases"] if "Test Cases" in wb.sheetnames else wb.worksheets[0]
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if header:
                cols = {str(h).strip(): i for i, h in enumerate(header)}
                idx_id = cols.get("ID")
                if idx_id is not None:
                    for row in rows:
                        cid = row[idx_id]
                        if cid is None:
                            continue
                        meta = {}
                        for col_name in ("Status", "Tanggal Run", "Environment"):
                            idx = cols.get(col_name)
                            val = row[idx] if idx is not None and idx < len(row) else None
                            if val not in (None, ""):
                                meta[col_name] = str(val).strip()
                        if meta:
                            existing[str(cid).strip()] = meta
            wb.close()
            return existing
        except Exception:
            pass  # XLSX tidak bisa dibaca → fallback ke tracker CSV

    csv_path = ROOT / "qa-test-cases-tracker.csv"
    if csv_path.exists():
        with csv_path.open(encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                cid = (row.get("ID") or "").strip()
                if not cid:
                    continue
                meta = {}
                for col_name in ("Status", "Tanggal Run", "Environment"):
                    val = (row.get(col_name) or "").strip()
                    if val:
                        meta[col_name] = val
                if meta:
                    existing[cid] = meta
    return existing


def enrich(records: list[dict],
           existing_metadata: dict[str, dict[str, str]] | None = None) -> list[dict]:
    existing = existing_metadata or {}
    for r in records:
        r["story"] = story_of(r["ref"])
        r["sprint"] = sprint_of(r["ref"])
        r["module"] = module_of(r)
        r["priority"] = SEVERITY_PRIORITY.get(r["sev"], "3 - Medium")
        r["type"] = "Regression" if r["id"].startswith("RG-") else "Functional"
        r["preconditions"] = precond_of(r)
        # Status/Tanggal Run/Environment yang sudah diisi QA per baris
        # dipertahankan; yang kosong jatuh ke default (Not Run / RUN_DATE /
        # ENVIRONMENT). Ini membuat regenerasi tidak menghapus data eksekusi.
        meta = existing.get(r["id"], {})
        r["status"] = meta.get("Status", "Not Run")
        # Saat --run-date CLI, paksa tanggal baru di semua baris;
        # tanpa override, pertahankan tanggal lama dari output sebelumnya.
        if _OVERRIDE_RUN_DATE:
            r["run_date"] = RUN_DATE_DEFAULT
        else:
            r["run_date"] = meta.get("Tanggal Run", RUN_DATE_DEFAULT)
        r["environment"] = meta.get("Environment", ENVIRONMENT_DEFAULT)
        # Pembuat test — default tim QA, bisa diedit per baris di Excel
        r["author"] = TEST_AUTHOR_DEFAULT
        # Link bug/defect — diisi QA saat ada kegagalan (bisa disinkronkan ke
        # kolom Defects di import TestRail). Kosong = belum ada defect.
        r["bug_link"] = ""
    return records


# ---------------------------------------------------------------------------
# Contoh run (--sample): status acak-deterministik sebagai referensi format
# ---------------------------------------------------------------------------
# status_id TestRail — SAMA dengan scripts/convert-results-to-testrail.py
SAMPLE_STATUS_IDS = {"Passed": 1, "Blocked": 2, "Retest": 4, "Failed": 5, "Skipped": 6}

# Seed status acak untuk --sample. TETAP — jangan ubah: contoh run harus
# bisa direproduksi persis kapan pun dijalankan ulang (nilai dicetak di
# output --sample, lihat juga QA Test Plan §8.1).
SAMPLE_SEED = 20260816

SAMPLE_FILES = {
    "tracker": ROOT / "qa-test-cases-sample-tracker.csv",
    "xlsx": ROOT / "qa-test-cases-sample.xlsx",
    "results": ROOT / "qa-test-results-sample-testrail.csv",
    "payload": ROOT / "qa-test-results-sample.json",
}


def random_statuses(records: list[dict], seed: int = SAMPLE_SEED) -> list[dict]:
    """Isi Status dengan nilai acak DETERMINISTIK (seed tetap) agar contoh run
    bisa direproduksi. Distribusi condong ke Passed, sisanya variasi status."""
    import random
    rng = random.Random(seed)
    pool = (["Passed"] * 60 + ["Failed"] * 10 + ["Blocked"] * 8
            + ["Retest"] * 7 + ["Skipped"] * 5 + ["Not Run"] * 10)
    for r in records:
        r["status"] = rng.choice(pool)
    return records


def build_sample_payload(records: list[dict]) -> dict:
    """Payload add_results_for_cases untuk contoh run. case_id memakai nomor
    urut (1..N) sebagai referensi FORMAT — konverter sesungguhnya butuh ID
    case TestRail yang valid (numerik atau via --mapping)."""
    results = []
    for i, r in enumerate(records, 1):
        sid = SAMPLE_STATUS_IDS.get(r["status"])
        if sid is None:  # Not Run (dan status lain yang tak bisa dikirim)
            continue
        results.append({"case_id": i, "status_id": sid})
    return {"results": results}


# ---------------------------------------------------------------------------
# Conditional formatting kolom Status (warna otomatis, ikut nilai sel)
# ---------------------------------------------------------------------------
def add_status_cf(ws, col: str, first_row: int, last_row: int) -> None:
    """Pasang conditional formatting di kolom Status: warna otomatis sesuai
    nilai sel (Passed hijau, Fail/Failed merah, Not Run abu-abu, Blocked/Retest
    amber, Skipped biru-abu). Berlaku untuk seluruh rentang data yang diberikan
    — warna ikut berubah jika QA mengedit nilai di Excel.
    """
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Font, PatternFill

    palette = [
        (("Passed",), "C6EFCE", "006100"),
        (("Fail", "Failed"), "FFC7CE", "9C0006"),
        (("Not Run",), "D9D9D9", "595959"),
        (("Blocked", "Retest"), "FFEB9C", "9C6500"),
        (("Skipped",), "DDEBF7", "1F4E78"),
    ]
    rng = f"{col}{first_row}:{col}{last_row}"
    for statuses, fill_hex, font_hex in palette:
        cond = " OR ".join(f'${col}{first_row}="{s}"' for s in statuses)
        if len(statuses) == 1:
            cond = f'${col}{first_row}="{statuses[0]}"'
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[cond],
                fill=PatternFill("solid", fgColor=fill_hex),
                font=Font(color=font_hex, bold=True),
                stopIfTrue=False,
            ),
        )


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------
def write_testrail_csv(records: list[dict]) -> Path:
    path = ROOT / "qa-test-cases-testrail.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        # lineterminator='\n': output LF deterministik lintas platform
        # (default csv module '\r\n' membuat byte berbeda di CI Linux vs
        # checkout Windows — lihat check-qa-sync.py).
        w = csv.writer(f, lineterminator="\n")
        w.writerow(["Case ID", "Title", "Section", "Type", "Priority", "Estimate",
                    "References", "Preconditions", "Steps", "Expected"])
        for r in records:
            w.writerow([r["id"], r["title"], r["module"], r["type"], r["priority"],
                        "", r["story"], r["preconditions"], r["steps"], r["expected"]])
    return path


def write_tracker_csv(records: list[dict], path: Path | None = None) -> Path:
    path = path or ROOT / "qa-test-cases-tracker.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        # lineterminator='\n' — output LF deterministik (lihat di atas).
        w = csv.writer(f, lineterminator="\n")
        w.writerow(["ID", "Modul", "Test Case", "Tipe", "Prioritas", "Severity",
                    "Story (AC)", "Sprint", "Status", "Tanggal Run", "Environment",
                    "Pembuat Test", "Link Bug",
                    "Langkah", "Hasil Diharapkan", "Preconditions"])
        for r in records:
            w.writerow([r["id"], r["module"], r["title"], r["type"], r["priority"],
                        r["sev"], r["story"], r["sprint"], r["status"],
                        r["run_date"], r["environment"], r["author"], r["bug_link"],
                        r["steps"], r["expected"], r["preconditions"]])
    return path


def write_xlsx(records: list[dict], path: Path | None = None) -> Path:
    from datetime import datetime, timezone
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # Pin timestamps agar XLSX byte-deterministik (hash stabil antar run)
    _epoch = datetime(2026, 1, 1, tzinfo=timezone.utc)
    wb.properties.created = _epoch
    wb.properties.modified = _epoch
    ws = wb.active
    ws.title = "Test Cases"

    headers = ["ID", "Modul", "Test Case", "Tipe", "Prioritas", "Severity",
               "Story (AC)", "Sprint", "Status", "Tanggal Run", "Environment",
               "Pembuat Test", "Link Bug",
               "Langkah", "Hasil Diharapkan", "Preconditions"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="2596BE")
    sev_fill = {
        "S1": PatternFill("solid", fgColor="FDE8E8"),
        "S2": PatternFill("solid", fgColor="FEF3C7"),
    }
    # Peringatan: baris S1 yang masih Not Run / Fail disorot merah
    warn_fill = PatternFill("solid", fgColor="FCA5A5")
    for col, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")

    for r in records:
        ws.append([r["id"], r["module"], r["title"], r["type"], r["priority"],
                   r["sev"], r["story"], r["sprint"], r["status"],
                   r["run_date"], r["environment"], r["author"], r["bug_link"],
                   r["steps"], r["expected"], r["preconditions"]])
        row = ws.max_row
        if r["sev"] in sev_fill:
            ws.cell(row=row, column=6).fill = sev_fill[r["sev"]]
        if r["sev"] == "S1" and r["status"] in OPEN_STATUSES:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = warn_fill
            c = ws.cell(row=row, column=9)  # Status
            c.font = Font(bold=True, color="9B1C1C")
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).alignment = Alignment(
                vertical="top", wrap_text=(col in (14, 15, 16)))

    # Warna otomatis kolom Status (Passed/Failed/Not Run, dll.)
    add_status_cf(ws, "I", 2, ws.max_row)

    widths = [14, 20, 38, 12, 14, 9, 12, 9, 11, 12, 20, 12, 14, 60, 60, 38]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # ---- Sheet ringkasan ----
    from collections import Counter
    ws2 = wb.create_sheet("Ringkasan")
    total = len(records)
    tc = [r for r in records if r["id"].startswith("TC-")]
    rg = [r for r in records if r["id"].startswith("RG-")]

    ws2.append(["Metrik", "Jumlah"])
    ws2.append(["Total test case", total])
    ws2.append(["  Test case fungsional (TC-*)", len(tc)])
    ws2.append(["  Skenario regresi (RG-*)", len(rg)])
    ws2.append([])
    ws2.append(["Per Modul", "Jumlah"])
    for mod, n in Counter(r["module"] for r in tc).most_common():
        if mod:
            ws2.append([mod, n])
    ws2.append([])
    ws2.append(["Per Severity", "Jumlah"])
    for sev, n in Counter(r["sev"] for r in tc).most_common():
        if sev:
            ws2.append([sev, n])
    ws2.append([])
    ws2.append(["Per Sprint", "Jumlah"])
    for sp, n in Counter(r["sprint"] for r in tc).most_common():
        if sp:
            ws2.append([sp, n])

    # ---- Peringatan S1 (Not Run / Fail) ----
    s1_all = [r for r in tc if r["sev"] == "S1"]
    s1_open = [r for r in s1_all if r["status"] in OPEN_STATUSES]
    warn_fill = PatternFill("solid", fgColor="FCA5A5")
    ws2.append([])
    ws2.append(["⚠️ PERINGATAN — Test case S1 yang belum lolos", ""])
    ws2.append([f"Total S1: {len(s1_all)} · Masih Not Run/Fail: {len(s1_open)}", ""])
    if s1_open:
        ws2.append(["ID", "Status"])
        s1_first_row = ws2.max_row + 1
        for r in s1_open:
            ws2.append([r["id"], r["status"]])
        # Warna otomatis kolom Status di daftar S1 — konsisten dengan sheet
        # Test Cases (Passed hijau / Fail-Failed merah / Not Run abu-abu / …)
        add_status_cf(ws2, "B", s1_first_row, ws2.max_row)
        ws2.append([])
        ws2.append(["Blokir rilis (release gate)", "YA — selesaikan semua S1 dulu"])
    else:
        ws2.append(["Semua S1 sudah dijalankan dan lolos", "✓"])
        ws2.append([])
        ws2.append(["Blokir rilis (release gate)", "TIDAK"])
    warn_row = ws2.max_row
    for col in (1, 2):
        c = ws2.cell(row=warn_row, column=col)
        c.fill = warn_fill
        c.font = Font(bold=True)
    # Sorot header peringatan
    warn_hdr_row = next(r[0].row for r in ws2.iter_rows(min_row=1, max_row=ws2.max_row) if r[0].value and str(r[0].value).startswith("⚠️"))
    for col in (1, 2):
        ws2.cell(row=warn_hdr_row, column=col).fill = warn_fill
        ws2.cell(row=warn_hdr_row, column=col).font = Font(bold=True, color="9B1C1C")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20
    for row in ws2.iter_rows(min_row=1, max_row=1):
        for c in row:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = header_fill

    path = path or ROOT / "qa-test-cases.xlsx"
    wb.save(path)
    _pin_xlsx_modified(path)
    return path


def run_payload_converter() -> None:
    """Jalankan otomatis konverter payload TestRail (convert-results-to-testrail.py)
    setelah CSV hasil diregenerasi.

    Status kosong / 'Not Run' / ID non-numerik tanpa --mapping dilewati
    (aturan konverter). Bila tidak ada result siap kirim, beri INFO tanpa
    menggagalkan generator — CSV hasil memang TEMPLATE yang diisi QA per run;
    begitu status terisi, regenerasi berikutnya otomatis menulis
    `qa-test-results-testrail.json`.
    """
    results_csv = ROOT / "qa-test-results-testrail.csv"
    out_json = ROOT / "qa-test-results-testrail.json"
    try:
        spec = importlib.util.spec_from_file_location(
            "convert_results_to_testrail",
            ROOT / "scripts" / "convert-results-to-testrail.py",
        )
        assert spec and spec.loader is not None
        conv = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(conv)
        results, stats = conv.convert(results_csv, {})
    except Exception as e:  # konverter rusak/jarang — jangan hentikan generator
        print(f"[WARN] Konverter payload tidak dijalankan: {e}")
        return
    if not results:
        print("[INFO] Payload TestRail kosong — belum ada status terisi yang bisa "
              "dikirim (atau butuh --mapping untuk ID non-numerik). "
              "JSON tidak ditulis.")
        return
    out_json.write_text(
        json.dumps({"results": results}, indent=2, ensure_ascii=False),
        encoding="utf-8")
    print(f"[OK] Payload TestRail: {out_json} ({len(results)} result siap kirim)")


def write_results_csv(records: list[dict], path: Path | None = None) -> Path:
    """Template hasil eksekusi per run — siap di-import balik ke TestRail
    sebagai result (Test Run → Import Results → CSV). Kolom mengikuti format
    importer TestRail: identifier case (Test Case ID) + field hasil.

    Status diisi QA per run: Passed / Failed / Blocked / Retest / Skipped
    (persis nama status TestRail). Kolom kosong lainnya opsional.
    """
    path = path or ROOT / "qa-test-results-testrail.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        # lineterminator='\n' — output LF deterministik (lihat di atas).
        w = csv.writer(f, lineterminator="\n")
        w.writerow(["Test Case ID", "Title", "Status", "Comment", "Elapsed",
                    "Version", "Defects", "Assignee"])
        for r in records:
            w.writerow([r["id"], r["title"], "", "", "", "", "", ""])
    return path


def write_results_template_xlsx(records: list[dict]) -> Path:
    """Workbook eksekusi per run: isi Status (dropdown), ringkasan otomatis
    via COUNTIF, dan release gate S1 (formula COUNTIFS).
    """
    from datetime import datetime, timezone
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    # Pin timestamps agar XLSX byte-deterministik (hash stabil antar run)
    _epoch = datetime(2026, 1, 1, tzinfo=timezone.utc)
    wb.properties.created = _epoch
    wb.properties.modified = _epoch
    ws = wb.active
    ws.title = "Hasil Run"

    # ---- Blok info run (diisi/ubah per run) ----
    ws.append(["RUN: Eksekusi Test — Appsheet Accounting Journal", ""])
    ws.append(["Nama Run", "Regresi Maret 2026"])
    ws.append(["Tanggal Run", RUN_DATE_DEFAULT])
    ws.append(["Environment", ENVIRONMENT_DEFAULT])
    ws.append(["Eksekutor", ""])
    ws.append([])

    header_fill = PatternFill("solid", fgColor="2596BE")
    warn_fill = PatternFill("solid", fgColor="FCA5A5")
    ok_fill = PatternFill("solid", fgColor="C6EFCE")

    headers = ["Test Case ID", "Modul", "Test Case", "Severity", "Sprint",
               "Status", "Comment", "Elapsed", "Defects", "Assignee"]
    ws.append(headers)
    header_row = ws.max_row  # 7
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=header_row, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")

    sev_fill = {"S1": PatternFill("solid", fgColor="FDE8E8"),
                "S2": PatternFill("solid", fgColor="FEF3C7")}
    first_data = header_row + 1
    for r in records:
        ws.append([r["id"], r["module"], r["title"], r["sev"], r["sprint"],
                   "", "", "", "", ""])
        row = ws.max_row
        if r["sev"] in sev_fill:
            ws.cell(row=row, column=4).fill = sev_fill[r["sev"]]
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).alignment = Alignment(
                vertical="top", wrap_text=(col in (3, 7)))
    last_data = ws.max_row

    # Dropdown status (persis nama status TestRail)
    dv = DataValidation(
        type="list",
        formula1='"Passed,Failed,Blocked,Retest,Skipped,Not Run"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(f"F{first_data}:F{last_data}")

    # Warna otomatis kolom Status — ikut nilai dropdown yang dipilih
    add_status_cf(ws, "F", first_data, last_data)

    # ---- Ringkasan otomatis (formula COUNTIF — dihitung Excel saat dibuka) ----
    ws.append([])
    ws.append(["RINGKASAN RUN", ""])
    rng = f"F{first_data}:F{last_data}"
    for i, status in enumerate(["Passed", "Failed", "Blocked", "Retest", "Skipped", "Not Run"], 1):
        ws.append([f"  {status}", f"=COUNTIF({rng},\"{status}\")"])
    ws.append(["  Total dijalankan (bukan Not Run)", f"=COUNTA({rng})-COUNTIF({rng},\"Not Run\")-COUNTIF({rng},\"\")"])
    ws.append(["  % Pass", f'=IF(COUNTA({rng})=0,0,COUNTIF({rng},\"Passed\")/COUNTA({rng}))'])
    ws.append([])
    ws.append(["RELEASE GATE (S1)", ""])
    ws.append(["  S1 belum lolos (Failed/Not Run)",
               f'=COUNTIFS(D{first_data}:D{last_data},"S1",F{first_data}:F{last_data},"<>Passed")-COUNTIFS(D{first_data}:D{last_data},"S1",F{first_data}:F{last_data},"",F{first_data}:F{last_data},"<>")'])
    ws.append(["  Blokir rilis jika S1 belum lolos",
               f'=IF(COUNTIFS(D{first_data}:D{last_data},"S1",F{first_data}:F{last_data},"<>Passed")>0,"YA — selesaikan S1 dulu","TIDAK")'])

    warn_hdr = next(r[0].row for r in ws.iter_rows(min_row=1, max_row=ws.max_row) if r[0].value and str(r[0].value).startswith("RELEASE GATE"))
    for col in (1, 2):
        ws.cell(row=warn_hdr, column=col).fill = warn_fill
        ws.cell(row=warn_hdr, column=col).font = Font(bold=True, color="9B1C1C")

    widths = [14, 20, 38, 9, 9, 11, 46, 9, 12, 14]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{last_data}"

    # ---- Sheet cara pakai ----
    ws2 = wb.create_sheet("Cara Pakai")
    lines = [
        ["Template hasil eksekusi test — import balik ke TestRail", ""],
        ["", ""],
        ["1. Isi kolom Status per case (dropdown): Passed / Failed / Blocked /", ""],
        ["   Retest / Skipped / Not Run (nama persis status TestRail).", ""],
        ["2. (Opsional) isi Comment, Elapsed (mis. \"5m\"), Defects (mis. TR-7), Assignee.", ""],
        ["3. Sheet 'Hasil Run' → ringkasan & release gate S1 terhitung otomatis.", ""],
        ["", ""],
        ["Import ke TestRail (cara A — UI):", ""],
        ["   Test Run → Import Results → pilih CSV. Format kolom importer:", ""],
        ["   Test Case ID (atau Case ID / Title) + Status + kolom hasil opsional.", ""],
        ["", ""],
        ["Import ke TestRail (cara B — API, add_results_for_cases):", ""],
        ["   Konversi CSV ini ke JSON payload dan POST ke run: ", ""],
        ["   curl -u user:key -H 'Content-Type: application/json' -F 'results=@res.json'", ""],
        ["   https://host/index.php?/api/v2/add_results_for_cases/{run_id}", ""],
        ["", ""],
        ["   status_id default: 1=Passed, 2=Blocked, 4=Retest, 5=Failed", ""],
        ["   (3=Untested tidak bisa dikirim sebagai hasil).", ""],
        ["", ""],
        ["Contoh payload satu hasil:", ""],
        ['   { "case_id": 100, "status_id": 5, "comment": "Gagal saat login",'],
        ['     "elapsed": "3m", "defects": "TR-7", "environment": "qa03" }'],
    ]
    for row in lines:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 95

    path = ROOT / "qa-test-results-template.xlsx"
    wb.save(path)
    _pin_xlsx_modified(path)
    return path


def main() -> None:
    # Output unicode aman di semua platform (mis. '—' di Windows cp1252)
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8")

    ap = argparse.ArgumentParser(
        description="Generate QA test case files (TestRail CSV + tracking CSV + XLSX)"
                    " dari `QA Test Plan - Accounting.md`.")
    ap.add_argument("--sample", action="store_true",
                    help="hasilkan CONTOH run: status acak-deterministik ke file "
                         "*-sample.* sebagai referensi format import "
                         "(file asli TIDAK disentuh)")
    ap.add_argument("--run-date", metavar="YYYY-MM-DD",
                    help="override tanggal run (default: env QA_RUN_DATE, lalu hari "
                         "ini). Dipakai juga oleh check-qa-sync di CI.")
    ap.add_argument("--environment",
                    help="override environment run (default: env QA_ENVIRONMENT, "
                         "lalu 'QA Local (mock API)')")
    ap.add_argument("--sample-seed", type=int, default=SAMPLE_SEED,
                    metavar="N",
                    help="seed untuk status acak di --sample (default: "
                         f"{SAMPLE_SEED}). Beda seed → distribusi berbeda, "
                         "tapi tetap deterministik — run ulang dengan seed "
                         "sama selalu menghasilkan status sama persis.")
    ap.add_argument("--update-baseline", action="store_true",
                    help="perbarui EXPECTED_TC/EXPECTED_RG di source code "
                         "ke jumlah aktual. Berguna saat QA Test Plan "
                         "bertambah/berkurang TC/RG.")
    args = ap.parse_args()

    # Terapkan override CLI ke nilai default (berlaku untuk seluruh output;
    # disimpan di module global agar dibaca enrich()/write_* saat dipanggil).
    global RUN_DATE_DEFAULT, ENVIRONMENT_DEFAULT, _OVERRIDE_RUN_DATE
    if args.run_date:
        RUN_DATE_DEFAULT = args.run_date
        _OVERRIDE_RUN_DATE = True
    if args.environment:
        ENVIRONMENT_DEFAULT = args.environment

    records = parse_tables()
    tc = [r for r in records if r["id"].startswith("TC-")]
    rg = [r for r in records if r["id"].startswith("RG-")]

    # --update-baseline: perbarui EXPECTED_TC/EXPECTED_RG di source code
    if args.update_baseline:
        _update_baseline(len(tc), len(rg))

    # Baseline lunak — bukan kontrak keras: jumlah yang berbeda dari EXPECTED_*
    # hanya memunculkan warning (penambahan TC ke plan tidak menghentikan run).
    for label, actual, expected in (("TC", len(tc), EXPECTED_TC),
                                    ("RG", len(rg), EXPECTED_RG)):
        if actual != expected:
            print(f"[WARN] Jumlah {label} {actual} != baseline {expected} — "
                  f"QA Test Plan bertambah/berkurang? Perbarui EXPECTED_{label} "
                  f"di scripts/generate-qa-test-cases.py bila perubahan ini disengaja.")
            if not args.update_baseline:
                print(f"  → Jalankan: python scripts/generate-qa-test-cases.py --update-baseline")

    if args.sample:
        import json
        sample_seed = args.sample_seed
        random_statuses(enrich(records), seed=sample_seed)  # tanpa load existing — status murni acak
        paths = [
            write_tracker_csv(records, SAMPLE_FILES["tracker"]),
            write_xlsx(records, SAMPLE_FILES["xlsx"]),
            write_results_csv(records, SAMPLE_FILES["results"]),
        ]
        SAMPLE_FILES["payload"].write_text(
            json.dumps(build_sample_payload(records), indent=2, ensure_ascii=False),
            encoding="utf-8")
        paths.append(SAMPLE_FILES["payload"])
        for p in paths:
            print(f"[OK] {p}  (contoh run — status acak, file asli tidak disentuh)")
        print(f"\nContoh run: {len(tc)} TC + {len(rg)} RG = {len(records)} test case dengan status terisi")
        print(f"         Seed status acak: {sample_seed} (tetap) — run ulang dengan seed sama menghasilkan status sama persis")
        return

    existing = load_existing_metadata()
    records = enrich(records, existing)
    carried = sum(1 for r in records if r["id"] in existing)
    if carried:
        print(f"[INFO] Mempertahankan status/metadata {carried} test case dari output sebelumnya")
    for f in (write_testrail_csv, write_tracker_csv, write_xlsx, write_results_csv, write_results_template_xlsx):
        print(f"[OK] {f(records)}")
    print(f"\nTotal: {len(tc)} TC + {len(rg)} RG = {len(records)} test case")
    # Otomatis lanjut ke konverter payload TestRail (bila ada status terisi)
    run_payload_converter()


if __name__ == "__main__":
    main()
