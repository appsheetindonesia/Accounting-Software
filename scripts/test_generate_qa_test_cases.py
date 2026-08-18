#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Unit test kecil untuk `load_existing_metadata` (scripts/generate-qa-test-cases.py).

Kasus yang diverifikasi:
  1. XLSX valid ada → metadata (Status + Tanggal Run + Environment) dibaca
     dari XLSX — menang atas tracker CSV (prioritas file terbaru).
  2. XLSX korup → fallback otomatis ke `qa-test-cases-tracker.csv`.
  3. Sel kosong diabaikan → kolom kosong TIDAK masuk metadata (saat enrich
     kolom tsb jatuh ke default), dan case dengan SEMUA sel kosong di-skip.

Menjalankan:  python -m unittest discover -s scripts -p "test_*.py"
"""
from __future__ import annotations

import csv
import importlib.util
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

# Muat modul generator tanpa mengeksekusi main() (import aman: hanya konstanta
# + definisi fungsi). ROOT di-patch per test agar tidak menyentuh repo asli.
_SPEC = importlib.util.spec_from_file_location(
    "generate_qa_test_cases",
    Path(__file__).resolve().parent / "generate-qa-test-cases.py",
)
gen = importlib.util.module_from_spec(_SPEC)
assert _SPEC.loader is not None
_SPEC.loader.exec_module(gen)


def _write_tracker(path: Path, rows: list[dict]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["ID", "Status", "Tanggal Run", "Environment"])
        w.writeheader()
        w.writerows(rows)


def _write_xlsx(path: Path, rows: list[dict]) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    ws.append(["ID", "Status", "Tanggal Run", "Environment"])
    for r in rows:
        ws.append([r.get("ID"), r.get("Status"), r.get("Tanggal Run"), r.get("Environment")])
    wb.save(path)
    wb.close()


class LoadExistingMetadataTest(unittest.TestCase):
    """Isolasi: ROOT dipatch ke direktori temp agar generator tidak membaca
    / menulis file output asli di repo."""

    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.root = Path(self._tmp.name)
        self.tracker = self.root / "qa-test-cases-tracker.csv"
        self.xlsx = self.root / "qa-test-cases.xlsx"
        patcher = patch.object(gen, "ROOT", self.root)
        patcher.start()
        self.addCleanup(patcher.stop)
        self.addCleanup(self._tmp.cleanup)

    def test_xlsx_menang_atas_csv(self) -> None:
        """XLSX + CSV sama-sama ada → metadata dari XLSX (prioritas)."""
        _write_xlsx(self.xlsx, [
            {"ID": "TC-A-01", "Status": "Passed", "Tanggal Run": "2026-08-20",
             "Environment": "QA Staging"},
        ])
        _write_tracker(self.tracker, [
            {"ID": "TC-A-01", "Status": "Failed", "Tanggal Run": "2026-08-01",
             "Environment": "Env CSV"},
        ])

        meta = gen.load_existing_metadata()

        self.assertEqual(meta["TC-A-01"]["Status"], "Passed")
        self.assertEqual(meta["TC-A-01"]["Tanggal Run"], "2026-08-20")
        self.assertEqual(meta["TC-A-01"]["Environment"], "QA Staging")

    def test_xlsx_korup_fallback_ke_csv(self) -> None:
        """XLSX ada tapi korup (bukan zip valid) → fallback ke tracker CSV."""
        self.xlsx.write_bytes(b"bukan zip xlsx yang valid \x00\x01" * 20)
        _write_tracker(self.tracker, [
            {"ID": "TC-B-01", "Status": "Failed", "Tanggal Run": "2026-08-10",
             "Environment": "Env CSV"},
        ])

        meta = gen.load_existing_metadata()

        self.assertEqual(meta["TC-B-01"]["Status"], "Failed")
        self.assertEqual(meta["TC-B-01"]["Tanggal Run"], "2026-08-10")
        self.assertEqual(meta["TC-B-01"]["Environment"], "Env CSV")

    def test_sel_kosong_diabaikan(self) -> None:
        """Kolom kosong tidak masuk metadata; case dengan semua sel kosong di-skip."""
        _write_xlsx(self.xlsx, [
            # Status terisi, tanggal/environment kosong → hanya Status yang ada
            {"ID": "TC-C-01", "Status": "Passed", "Tanggal Run": "", "Environment": ""},
            # Hanya Tanggal Run terisi → Status tidak ada (enrich → 'Not Run')
            {"ID": "TC-C-02", "Status": "", "Tanggal Run": "2026-08-15", "Environment": ""},
            # Semua kosong → case tidak muncul sama sekali
            {"ID": "TC-C-03", "Status": None, "Tanggal Run": None, "Environment": None},
        ])

        meta = gen.load_existing_metadata()

        self.assertEqual(meta["TC-C-01"]["Status"], "Passed")
        self.assertNotIn("Tanggal Run", meta["TC-C-01"])
        self.assertNotIn("Environment", meta["TC-C-01"])

        self.assertEqual(meta["TC-C-02"]["Tanggal Run"], "2026-08-15")
        self.assertNotIn("Status", meta["TC-C-02"])

        self.assertNotIn("TC-C-03", meta)

    def test_tanpa_output_sebelumnya_return_kosong(self) -> None:
        """Tidak ada XLSX maupun CSV → dict kosong (semua jatuh ke default)."""
        self.assertEqual(gen.load_existing_metadata(), {})


class SampleDeterminismTest(unittest.TestCase):
    """Jalankan --sample 2x via subprocess → bandingkan hash CSV/JSON
    dan konten sel XLSX. Membuktikan output generator deterministik.
    """

    def _run_sample(self) -> dict[str, bytes]:
        """Jalankan generator --sample via subprocess, kembalikan {name: bytes}."""
        import subprocess
        result = subprocess.run(
            [sys.executable, str(Path(__file__).resolve().parent / "generate-qa-test-cases.py"), "--sample"],
            capture_output=True, text=True, timeout=30,
            cwd=str(Path(__file__).resolve().parent.parent),
        )
        self.assertEqual(result.returncode, 0,
                         f"Generator gagal:\nstdout={result.stdout}\nstderr={result.stderr}")
        # Kumpulkan file sample
        repo = Path(__file__).resolve().parent.parent
        files = {}
        for name in ("qa-test-cases-sample-tracker.csv",
                     "qa-test-cases-sample.xlsx",
                     "qa-test-results-sample-testrail.csv",
                     "qa-test-results-sample.json"):
            p = repo / name
            if p.exists():
                files[name] = p.read_bytes()
        return files

    def test_csv_hash_stabil(self) -> None:
        """Hash CSV identik antar 2 run."""
        import hashlib
        run1 = self._run_sample()
        run2 = self._run_sample()
        for name in run1:
            if name.endswith(".csv"):
                h1 = hashlib.md5(run1[name]).hexdigest()
                h2 = hashlib.md5(run2[name]).hexdigest()
                self.assertEqual(h1, h2, f"Hash {name} berbeda: {h1} vs {h2}")

    def test_json_hash_stabil(self) -> None:
        """Hash JSON identik antar 2 run."""
        import hashlib
        run1 = self._run_sample()
        run2 = self._run_sample()
        for name in run1:
            if name.endswith(".json"):
                h1 = hashlib.md5(run1[name]).hexdigest()
                h2 = hashlib.md5(run2[name]).hexdigest()
                self.assertEqual(h1, h2, f"Hash {name} berbeda: {h1} vs {h2}")

    def test_xlsx_cell_content_stabil(self) -> None:
        """Konten sel XLSX identik antar 2 run (bukan hanya hash file)."""
        from openpyxl import load_workbook
        repo = Path(__file__).resolve().parent.parent
        # Run 1
        self._run_sample()
        xlsx_files_1 = sorted(repo.glob("qa-test-cases-sample.xlsx"))
        contents_1 = []
        for xf in xlsx_files_1:
            wb = load_workbook(xf)
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    contents_1.append((xf.name, ws.title, row))
            wb.close()
        # Run 2
        self._run_sample()
        xlsx_files_2 = sorted(repo.glob("qa-test-cases-sample.xlsx"))
        contents_2 = []
        for xf in xlsx_files_2:
            wb = load_workbook(xf)
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    contents_2.append((xf.name, ws.title, row))
            wb.close()
        # Bandingkan
        self.assertEqual(len(contents_1), len(contents_2),
                         "Jumlah sel berbeda antar run")
        for i, (c1, c2) in enumerate(zip(contents_1, contents_2)):
            self.assertEqual(c1, c2, f"Sel #{i} berbeda: {c1} vs {c2}")

    def test_xlsx_hash_stabil(self) -> None:
        """Hash file XLSX identik antar 2 run (byte-deterministik)."""
        import hashlib
        run1 = self._run_sample()
        run2 = self._run_sample()
        for name in run1:
            if name.endswith(".xlsx"):
                h1 = hashlib.md5(run1[name]).hexdigest()
                h2 = hashlib.md5(run2[name]).hexdigest()
                self.assertEqual(h1, h2, f"Hash {name} berbeda: {h1} vs {h2}")

    def test_all_4_artifacts_stabil(self) -> None:
        """Semua 4 artefak (2 CSV + 1 JSON + 1 XLSX) hash stabil."""
        import hashlib
        run1 = self._run_sample()
        run2 = self._run_sample()
        self.assertEqual(set(run1.keys()), set(run2.keys()),
                         "Jenis file berbeda antar run")
        for name in run1:
            h1 = hashlib.md5(run1[name]).hexdigest()
            h2 = hashlib.md5(run2[name]).hexdigest()
            self.assertEqual(h1, h2, f"Hash {name} berbeda: {h1} vs {h2}")


if __name__ == "__main__":
    unittest.main()
