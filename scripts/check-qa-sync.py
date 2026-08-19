#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verifikasi artefak QA selalu sinkron dengan `QA Test Plan - Accounting.md`.

Alur:
  1. Backup file output yang sedang ter-commit (state saat ini).
  2. Baca 'Tanggal Run' dari baris pertama `qa-test-cases-tracker.csv`
     yang ter-commit — ini jadi acuan tanggal.
  3. Jalankan generator (scripts/generate-qa-test-cases.py) dengan env
     QA_RUN_DATE di-override ke tanggal dari tracker. Ini memastikan
     generator mereproduksi artefak dengan tanggal yang SAMA persis
     dengan yang sudah di-commit — jadi pergantian hari TIDAK dianggap
     sebagai perubahan.
  4. Bandingkan KONTEN:
       - CSV        → identik SETELAH normalisasi line ending (safety net —
                      `.gitattributes` (`*.csv text eol=lf`) sudah menjamin
                      checkout selalu LF, jadi di alur normal byte mentah
                      sudah cocok; normalisasi tetap menjaga perbandingan
                      tahan terhadap editor yang menulis CRLF ke disk).
       - XLSX       → nilai sel per sheet + formula conditional formatting
                      (byte XLSX TIDAK dibandingkan: openpyxl menulis timestamp
                      zip non-deterministik, jadi dua save konten-identik
                      menghasilkan byte berbeda).

Keluar dengan exit code 0 jika sinkron, 1 jika ada perbedaan (untuk CI gate).

Menjalankan:  python scripts/check-qa-sync.py
"""
from __future__ import annotations

import csv
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
GENERATOR = ROOT / "scripts" / "generate-qa-test-cases.py"

OUTPUT_FILES = [
    "qa-test-cases-testrail.csv",
    "qa-test-cases-tracker.csv",
    "qa-test-cases.xlsx",
    "qa-test-results-testrail.csv",
    "qa-test-results-template.xlsx",
]


def normalize_line_endings(b: bytes) -> bytes:
    """Normalisasi line ending → LF. KEBUTUHAN UTAMA sudah hilang: `.gitattributes`
    (`*.csv text eol=lf`) menjamin checkout selalu LF di semua platform, jadi byte
    yang ditulis generator (LF) cocok langsung dengan disk. Fungsi ini tetap ada
    sebagai safety net — editor yang menyimpan CRLF ke disk (mis. Excel lama)
    tidak boleh membuat gate salah-tolak. CRLF dan CR (Mac lama) keduanya dianggap
    LF agar perbandingan hanya melihat perbedaan konten, bukan gaya line ending.
    Diuji di scripts/test_check_qa_sync.py."""
    return b.replace(b"\r\n", b"\n").replace(b"\r", b"\n")


def committed_run_date() -> str:
    """Ambil 'Tanggal Run' dari tracker yang ter-commit (baris data pertama).

    Dipakai sebagai acuan tanggal oleh check-qa-sync: generator dijalankan
    dengan env QA_RUN_DATE di-override ke tanggal ini, sehingga pergantian
    hari di runner CI TIDAK menyebabkan artefak dianggap berubah.
    """
    tracker = ROOT / "qa-test-cases-tracker.csv"
    if tracker.exists():
        with tracker.open(encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                d = (row.get("Tanggal Run") or "").strip()
                if d:
                    return d
    from datetime import date
    return date.today().isoformat()


def sheet_snapshot(path: Path) -> dict[str, tuple]:
    """Snapshot konten workbook: nilai tiap sel (values_only) + formula CF."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=False, data_only=False)
    snap: dict[str, tuple] = {}
    for ws in wb.worksheets:
        values = tuple(tuple(row) for row in ws.iter_rows(values_only=True))
        cf = tuple(sorted(
            (str(rng.sqref), tuple(r.formula for r in rng.rules))
            for rng in ws.conditional_formatting
        ))
        snap[ws.title] = (values, cf)
    wb.close()
    return snap


def main() -> None:
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8")

    run_date = committed_run_date()
    print(f"[INFO] Tanggal run acuan (dari tracker ter-commit): {run_date}")

    # 1. Backup state saat ini
    with tempfile.TemporaryDirectory() as tmp:
        backup = Path(tmp)
        for name in OUTPUT_FILES:
            src = ROOT / name
            if src.exists():
                shutil.copy2(src, backup / name)

        # 2. Regenerate dengan tanggal ter-pin (override env QA_RUN_DATE
        #    agar generator memakai tanggal dari tracker ter-commit, bukan
        #    hari ini — ini kunci agar pergantian hari tidak false-positive)
        env = dict(os.environ, QA_RUN_DATE=run_date)
        proc = subprocess.run([sys.executable, str(GENERATOR)],
                              cwd=ROOT, env=env, capture_output=True, text=True)
        if proc.returncode != 0:
            print(proc.stdout, proc.stderr, file=sys.stderr)
            raise SystemExit(f"[ERROR] Generator gagal dijalankan (exit {proc.returncode})")

        # 3. Bandingkan konten
        diffs: list[str] = []
        for name in OUTPUT_FILES:
            orig = backup / name
            new = ROOT / name
            if not orig.exists():
                diffs.append(f"{name}: tidak ada di state sebelumnya")
                continue
            if name.endswith(".csv"):
                # Normalisasi line ending (lihat normalize_line_endings)
                if (normalize_line_endings(orig.read_bytes())
                        != normalize_line_endings(new.read_bytes())):
                    diffs.append(f"{name}: isi CSV berbeda — QA Test Plan berubah "
                                 f"tanpa regenerasi")
            else:  # XLSX — bandingkan nilai + CF, bukan byte
                o, n = sheet_snapshot(orig), sheet_snapshot(new)
                if set(o) != set(n):
                    diffs.append(f"{name}: daftar sheet berbeda")
                else:
                    for sheet in o:
                        if o[sheet] != n[sheet]:
                            diffs.append(f"{name} [{sheet}]: isi/format berbeda")

        # 3b. Pulihkan state semula (regenerate menimpa file di tempat;
        #     byte XLSX non-deterministik — jangan biarkan working tree berubah)
        for name in OUTPUT_FILES:
            orig = backup / name
            if orig.exists():
                shutil.copy2(orig, ROOT / name)

        # 4. Laporan
        if diffs:
            print("[FAIL] Artefak QA TIDAK sinkron dengan QA Test Plan:")
            for d in diffs[:10]:
                print(f"  - {d}")
            if len(diffs) > 10:
                print(f"  … dan {len(diffs) - 10} perbedaan lain")
            print("\nPerbaiki: jalankan 'python scripts/generate-qa-test-cases.py' "
                  "dan commit hasilnya bersama perubahan QA Test Plan.")
            raise SystemExit(1)

        print("[OK] Artefak QA sinkron dengan QA Test Plan "
              f"({len(OUTPUT_FILES)} file, tanggal run {run_date})")


if __name__ == "__main__":
    main()
